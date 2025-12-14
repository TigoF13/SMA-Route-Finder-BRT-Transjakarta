import os
import json
from django.shortcuts import render
from django.conf import settings
from django.http import JsonResponse
from .forms import RouteForm  # Pastikan ini diimpor
from .pulp_solver.pulp import (
    build_transport_graph_with_costs,
    find_route_with_pulp_weighted,
    calculate_base_price,
    haversine
)
from .sma_solver.sma import calculate_path_fitness
from .sma_module import run_sma
from datetime import datetime, time
import pytz # Import pytz
import openpyxl # <-- Impor baru
from openpyxl.utils import get_column_letter # <-- Impor baru

# --- Cache Koordinat Halte (BARU) ---
HALTE_COORDS_CACHE = {}
try:
    # Gunakan path yang sama dengan cache nama halte
    nodes_path_global = os.path.join(settings.BASE_DIR, "myapp", "static", "data", "cleaned_nodes_new.geojson")
    with open(nodes_path_global, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Buat map unik {nama_halte: (lon, lat)}
    halte_coords = {}
    for feat in data["features"]:
        props = feat.get("properties", {})
        geom = feat.get("geometry", {})
        if "name" in props and geom.get("type") == "Point":
            name = props["name"]
            coords = geom.get("coordinates") # [lon, lat]
            # Pastikan nama unik dan koordinat valid
            if name not in halte_coords and coords and len(coords) == 2:
                halte_coords[name] = (coords[0], coords[1]) # Simpan (lon, lat)
                
    HALTE_COORDS_CACHE = halte_coords
    print(f"INFO: Berhasil memuat {len(HALTE_COORDS_CACHE)} koordinat halte untuk cache.")
except Exception as e:
    print(f"WARNING: Gagal memuat cache koordinat halte: {e}")
# ---------------------------------------------

# --- Cache Halte (INI YANG KEMUNGKINAN HILANG) ---
HALTE_NAMES_CACHE = []
try:
    nodes_path_global = os.path.join(settings.BASE_DIR, "myapp", "static", "data", "cleaned_nodes_new.geojson")
    with open(nodes_path_global, "r", encoding="utf-8") as f:
        data = json.load(f)
    HALTE_NAMES_CACHE = sorted(list(set([
        feat["properties"]["name"] for feat in data["features"] if "name" in feat["properties"]
    ])))
    print(f"INFO: Berhasil memuat {len(HALTE_NAMES_CACHE)} nama halte untuk autocomplete.")
except Exception as e:
    print(f"WARNING: Gagal memuat cache nama halte: {e}")
# ---------------------------------------------

EVAL_FILE_PATH = os.path.join(settings.BASE_DIR, "evaluation_log_skripsi.xlsx")

# Tentukan header untuk file Excel sesuai permintaan Anda
HEADER_ROW = [
    "Timestamp", "Halte Asal", "Halte Tujuan", "Preferensi Rute", "Waktu Keberangkatan",
    "Mode", "Param: Speed (km/h)", "Param: Wait (min)", "Param: Dwell (s)", 
    "Param: Delay (s/km)", "Waktu Komputasi MILP (s)", "Waktu Komputasi SMA (s)",
    "Nilai Fitness (Z_optimal)", "Nilai Fitness (Z_SMA)", "Optimality Gap (%)"
]

def append_to_excel_log(data_row):
    """
    Menambahkan satu baris data ke file evaluation_log_skripsi.xlsx.
    Membuat file dan header jika belum ada (sesuai permintaan Anda).
    """
    try:
        # Coba buka workbook yang ada
        try:
            wb = openpyxl.load_workbook(EVAL_FILE_PATH)
            ws = wb.active
        except FileNotFoundError:
            # File tidak ada, buat baru
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(HEADER_ROW) # Tambahkan header
            # Sesuaikan lebar kolom agar mudah dibaca
            for i, cell in enumerate(ws[1], 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = 22

        # Tambahkan baris data baru
        ws.append(data_row)
        
        # Simpan file
        wb.save(EVAL_FILE_PATH)
        print(f"INFO: Berhasil menyimpan log evaluasi ke {EVAL_FILE_PATH}")

    except Exception as e:
        # Jika gagal (misal: file terkunci/sedang Anda buka), 
        # cetak error ke terminal agar server tidak crash.
        print(f"WARNING: GAGAL menyimpan log Excel! Error: {e}")
        print(f"DATA LOG GAGAL TERSIMPAN: {data_row}")

# --------------------------------------------------

def get_walkable_neighbors(origin_stop_name, max_distance_km=1.0, walk_speed_kmh=5):
    """
    Menemukan halte tetangga dalam jarak berjalan kaki dari cache.
    Mengembalikan list of tuples: (nama_halte, waktu_jalan_menit, jarak_km).
    """
    neighbors = []
    if origin_stop_name not in HALTE_COORDS_CACHE:
        return [] # Halte asal tidak ada di cache koordinat

    origin_lon, origin_lat = HALTE_COORDS_CACHE[origin_stop_name]

    for stop_name, (lon, lat) in HALTE_COORDS_CACHE.items():
        if stop_name == origin_stop_name:
            continue
        
        # Hitung jarak
        dist_km = haversine(origin_lon, origin_lat, lon, lat)
        
        # Jika dalam radius, tambahkan sebagai kandidat
        if dist_km <= max_distance_km:
            walk_time_minutes = (dist_km / walk_speed_kmh) * 60
            neighbors.append((stop_name, walk_time_minutes, dist_km))
            
    return neighbors

def run_single_optimization(start_stop, end_stop, preferensi, jam_berangkat_str, metode_solver, G, stop_map, weights_dict, nodes_path, dynamic_params={}):
    """
    Refaktor dari logika solver Anda yang ada di 'index' view.
    Ini menjalankan satu skenario optimasi (MILP atau SMA).
    """
    print(f"INFO: Menjalankan optimasi untuk {start_stop} -> {end_stop} ({metode_solver})")
    
    dynamic_speed = dynamic_params.get('speed')
    dynamic_wait = dynamic_params.get('wait')
    
    dynamic_dwell = dynamic_params.get('dwell')
    dynamic_delay = dynamic_params.get('delay_km')
    
    if metode_solver == "milp":
        # Gunakan fungsi dari pulp.py
        hasil_rute = find_route_with_pulp_weighted(G, stop_map, start_stop, end_stop, weights=weights_dict, nodes_file=nodes_path)
    else: # sma
        # Gunakan fungsi dari sma_module.py
        hasil_rute = run_sma(start_stop, end_stop, preferensi_input=preferensi, waktu_keberangkatan=jam_berangkat_str, dynamic_speed=dynamic_speed, dynamic_wait=dynamic_wait, dynamic_dwell=dynamic_dwell, dynamic_delay=dynamic_delay)
    
    # Kembalikan hasil mentah dari solver
    return hasil_rute

# --- Fungsi Autocomplete (Tidak berubah) ---
def get_halte_list(request):
    query = request.GET.get("q", "").strip().lower()
    if not query or not HALTE_NAMES_CACHE: return JsonResponse([], safe=False)
    starts_with = [h for h in HALTE_NAMES_CACHE if h.lower().startswith(query)]
    contains_word = [
        h for h in HALTE_NAMES_CACHE
        if not h.lower().startswith(query) and any(word.startswith(query) for word in h.lower().split())
    ]
    filtered = (starts_with + contains_word)[:10]
    return JsonResponse(filtered, safe=False)

# --- Fungsi Index UTAMA (Dengan Persist Form & Waktu Default) ---
def index(request):
    hasil = {}
    form = None 

    if request.method == "POST":
        form = RouteForm(request.POST) # <-- Inisialisasi form dengan data POST
        if form.is_valid():
            # Proses data HANYA jika form valid
            halte_asal = form.cleaned_data["halte_asal"]
            halte_tujuan = form.cleaned_data["halte_tujuan"]
            preferensi = form.cleaned_data["preferensi"]
            
            # --- DAPATKAN OBJEK WAKTU ---
            jam_berangkat_obj = form.cleaned_data["jam_berangkat"] # Ini adalah objek datetime.time
            jam_berangkat = jam_berangkat_obj.strftime("%H:%M") # Ini adalah string
            
            metode_solver = form.cleaned_data["metode_solver"]

            # --- PENANGANAN ERROR AWAL ---
            if halte_asal == halte_tujuan:
                hasil = {"error": f"Halte Asal dan Tujuan tidak boleh sama ('{halte_asal}')."}
            elif HALTE_NAMES_CACHE and (halte_asal not in HALTE_NAMES_CACHE):
                hasil = {"error": f"Halte Asal '{halte_asal}' tidak ditemukan dalam data."}
            elif HALTE_NAMES_CACHE and (halte_tujuan not in HALTE_NAMES_CACHE):
                hasil = {"error": f"Halte Tujuan '{halte_tujuan}' tidak ditemukan dalam data."}
            else:
                # --- JIKA LOLOS, LANJUTKAN PROSES ---
                try:
                    
                    # --- LOGIKA BARU: TENTUKAN PARAMETER DINAMIS ---
                    
                    jam_sibuk_pagi_start = time(7, 0) # 07:00
                    jam_sibuk_pagi_end = time(9, 0)   # 09:00
                    jam_sibuk_sore_start = time(17, 0) # 17:00
                    jam_sibuk_sore_end = time(19, 0)   # 19:00

                    is_jam_sibuk = (jam_sibuk_pagi_start <= jam_berangkat_obj <= jam_sibuk_pagi_end) or \
                                  (jam_sibuk_sore_start <= jam_berangkat_obj <= jam_sibuk_sore_end)

                    if is_jam_sibuk:
                        print("INFO: Mode JAM SIBUK terdeteksi.")
                        param_speed = 15     # Kecepatan rata-rata lambat (15 km/jam)
                        param_wait = 5       # Waktu tunggu (5 menit)
                        param_dwell = 60     # Henti di halte (60 detik)
                        param_delay_km = 45  # Penalti macet per km (45 detik)
                    else:
                        print("INFO: Mode JAM NORMAL terdeteksi.")
                        param_speed = 20     # Kecepatan rata-rata (20 km/jam)
                        param_wait = 10      # Waktu tunggu (10 menit)
                        param_dwell = 45     # Henti di halte (45 detik)
                        param_delay_km = 30  # Penalti macet per km (30 detik)
                        
                    dynamic_params = {
                        "speed": param_speed, 
                        "wait": param_wait,
                        "dwell": param_dwell,
                        "delay_km": param_delay_km
                    }

                    # --- Persiapan Variabel Umum ---
                    if preferensi == "min_transit":
                        preferensi_label = "🔄 Minim Transit"
                        weights_dict = {"waktu": 0.1, "biaya": 0.1, "transit": 0.8}
                    elif preferensi == "cepat":
                        preferensi_label = "⚡ Paling Cepat"
                        weights_dict = {"waktu": 0.8, "biaya": 0.1, "transit": 0.1}
                    else: # efisien
                        preferensi_label = "🚍 Efisien & Seimbang"
                        weights_dict = {"waktu": 0.45, "biaya": 0.1, "transit": 0.45}

                    data_dir = os.path.join(settings.BASE_DIR, "myapp", "static", "data")
                    nodes_path = os.path.join(data_dir, "cleaned_nodes_new.geojson")
                    edges_path = os.path.join(data_dir, "transjakarta_edges.geojson")

                    print("\n--- Memulai Pencarian Rute Multi-Origin ---")
                    
                    # --- MODIFIKASI: Gunakan parameter dinamis saat membangun Graf Utama ---
                    print(f"INFO (Utama): Membangun graf dengan Speed={param_speed}, Wait={param_wait}, Dwell={param_dwell}, Delay={param_delay_km}")
                    
                    G, stop_map = build_transport_graph_with_costs(
                        nodes_path, 
                        edges_path,
                        avg_speed_kmh=param_speed,
                        avg_transfer_min=param_wait,
                        stop_dwell_time_seconds=param_dwell,
                        delay_per_km_seconds=param_delay_km
                    )
                    
                    if G is None:
                        hasil = {"error": "Gagal membangun graf dari file GeoJSON."}
                    else:
                        # --- 1. Dapatkan Kandidat Halte Asal (LOGIKA BARU) ---
                        origins_to_test = [
                            (halte_asal, 0.0, 0.0) # (nama_halte, waktu_jalan_menit, jarak_km)
                        ]
                        neighbors = get_walkable_neighbors(halte_asal, max_distance_km=1.0, walk_speed_kmh=5) 
                        origins_to_test.extend(neighbors)
                        
                        print(f"INFO: Menguji {len(origins_to_test)} kandidat halte asal: {[o[0] for o in origins_to_test]}")

                        # --- 2. Loop Optimasi (LOGIKA BARU) ---
                        best_result_payload = None
                        min_total_time = float('inf')

                        for start_stop, walk_time_min, walk_dist_km in origins_to_test:
                            
                            hasil_rute = run_single_optimization(
                                start_stop, halte_tujuan, preferensi, jam_berangkat, 
                                metode_solver, G, stop_map, weights_dict, nodes_path,
                                dynamic_params=dynamic_params 
                            )

                            if "error" in hasil_rute:
                                print(f"WARN: Skenario {start_stop} -> {halte_tujuan} gagal: {hasil_rute['error']}")
                                continue

                            bus_time_min = hasil_rute.get('waktu_tempuh_menit', 0)
                            wait_time = param_wait if bus_time_min > 0 else 0.0
                            total_time = walk_time_min + wait_time + bus_time_min

                            print(f"INFO: Skenario {start_stop}: Jalan {walk_time_min:.1f} m + Tunggu {wait_time:.1f} m + Bus {bus_time_min:.1f} m = Total {total_time:.1f} m")

                            if total_time < min_total_time:
                                min_total_time = total_time
                                best_result_payload = {
                                    "hasil_rute": hasil_rute, 
                                    "walk_time_min": walk_time_min,
                                    "walk_dist_km": walk_dist_km,
                                    "start_halte": start_stop, 
                                    "total_time_min": total_time,
                                    "initial_wait_min": wait_time 
                                }

                        # --- 3. Format Hasil Terbaik (LOGIKA BARU DENGAN SANITY CHECK) ---
                        if best_result_payload is None:
                            hasil = {"error": f"Tidak ditemukan rute dari {halte_asal} (atau halte terdekat) ke {halte_tujuan}."}
                        else:
                            rute_terbaik = best_result_payload["hasil_rute"]
                            start_halte = best_result_payload["start_halte"]
                            walk_time = best_result_payload["walk_time_min"]
                            walk_dist_km = best_result_payload["walk_dist_km"]
                            total_time = best_result_payload["total_time_min"]
                            bus_time = rute_terbaik.get('waktu_tempuh_menit', 0)
                            initial_wait = best_result_payload.get('initial_wait_min', 0)
                            is_walk_scenario = walk_dist_km > 0

                            if is_walk_scenario:
                                path_nodes = rute_terbaik.get("path_nodes", [])
                                time_to_origin_halte_min = 0.0
                                origin_halte_found_in_path = False
                                temp_time_hours = 0.0
                                
                                for i in range(len(path_nodes) - 1):
                                    u, v = path_nodes[i], path_nodes[i+1]
                                    if G.has_edge(u, v):
                                        temp_time_hours += G.get_edge_data(u, v).get('Waktuij', 0)
                                    else:
                                        temp_time_hours = 0.0 
                                        break 

                                    if v[0] == halte_asal:
                                        time_to_origin_halte_min = temp_time_hours * 60
                                        origin_halte_found_in_path = True
                                        break
                                        
                                if origin_halte_found_in_path and temp_time_hours > 0:
                                    print(f"INFO: SANITY CHECK! Rute {start_halte} melewati {halte_asal}. Mengkoreksi hasil.")
                                    bus_time = bus_time - time_to_origin_halte_min
                                    walk_time = 0.0
                                    walk_dist_km = 0.0
                                    is_walk_scenario = False
                                    start_halte = halte_asal 
                                    total_time = initial_wait + bus_time
                            
                            hasil = {
                                "preferensi_label": preferensi_label + f" ({metode_solver.upper()})",
                                "waktu_normal_fmt": f"{round(bus_time)} menit",
                                "waktu_macet_fmt": f"{round(bus_time * 1.8)} menit", 
                                "biaya_fmt": f"Rp {calculate_base_price(jam_berangkat):,}",
                                "jumlah_transit": rute_terbaik.get("jumlah_transit", 0),
                                "jam_berangkat": jam_berangkat,
                                "detailed_journey": rute_terbaik.get("detailed_journey", []),
                                "path_coords": rute_terbaik.get("path_coords", []),
                                "waktu_total_fmt": f"{round(total_time)} menit",
                                "initial_wait_min": round(initial_wait), 
                                "origin_info": {
                                    "is_walk": is_walk_scenario, 
                                    "start_halte": start_halte, 
                                    "walk_time_min": round(walk_time), 
                                    "walk_dist_km_fmt": f"{walk_dist_km:.1f}", 
                                    "original_start": halte_asal 
                                }
                            }

                        # --- 4. JALANKAN EVALUASI PERBANDINGAN DAN SIMPAN KE EXCEL ---
                        # --- BLOK INI TELAH DIUBAH ---
                        
                        print("\n--- Menjalankan Evaluasi Perbandingan (disimpan ke Excel) ---")
                        
                        start_time_milp = datetime.now()
                        eval_milp = find_route_with_pulp_weighted(G, stop_map, halte_asal, halte_tujuan, weights=weights_dict, nodes_file=nodes_path)
                        time_milp = (datetime.now() - start_time_milp).total_seconds()
                        
                        start_time_sma = datetime.now()
                        eval_sma = run_sma(halte_asal, halte_tujuan, 
                                           preferensi_input=preferensi, 
                                           waktu_keberangkatan=jam_berangkat,
                                           dynamic_speed=param_speed,
                                           dynamic_wait=param_wait,
                                           dynamic_dwell=param_dwell,
                                           dynamic_delay=param_delay_km)
                        time_sma = (datetime.now() - start_time_sma).total_seconds()
                        
                        Z_optimal = float('inf')
                        if "path_nodes" in eval_milp and "error" not in eval_milp:
                            Z_optimal = calculate_path_fitness(G, eval_milp["path_nodes"], weights_dict)
                        
                        Z_SMA = eval_sma.get('objective_cost', float('inf'))
                        gap = None
                        
                        if Z_optimal not in [0, float('inf')] and Z_SMA != float('inf'):
                            gap = ((Z_SMA - Z_optimal) / Z_optimal) * 100
                        
                        # Siapkan baris data untuk Excel
                        log_data_row = [
                            datetime.now(pytz.timezone('Asia/Jakarta')).strftime("%Y-%m-%d %H:%M:%S"),
                            halte_asal,
                            halte_tujuan,
                            preferensi,
                            jam_berangkat,
                            "JAM SIBUK" if is_jam_sibuk else "JAM NORMAL",
                            param_speed,
                            param_wait,
                            param_dwell,
                            param_delay_km,
                            round(time_milp, 2),
                            round(time_sma, 2),
                            round(Z_optimal, 4) if Z_optimal != float('inf') else "N/A",
                            round(Z_SMA, 4) if Z_SMA != float('inf') else "N/A",
                            round(gap, 2) if gap is not None else "N/A"
                        ]
                        
                        # Panggil helper untuk menyimpan ke Excel
                        append_to_excel_log(log_data_row)
                        
                        # --- AKHIR BLOK EVALUASI ---

                except Exception as e:
                    hasil = {"error": f"Terjadi kesalahan internal: {str(e)}"}
        else:
            # Jika form POST tidak valid
            hasil = {"error": "Input tidak valid. Silakan periksa kembali isian form."}
            # 'form' sudah berisi data POST dan errornya

    # --- Jika BUKAN POST (GET Request) atau form belum dibuat setelah POST gagal validasi ---
    if form is None or not form.is_bound: # Cek jika form belum di-bind (GET atau POST gagal)
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        now_jakarta = datetime.now(jakarta_tz)
        current_time_str = now_jakarta.strftime('%H:%M')
        form = RouteForm(initial={'jam_berangkat': current_time_str})
        if request.method == "GET": hasil = {} # Pastikan hasil kosong hanya untuk GET

    # Render template, kirim 'form' (kosong/awal atau terisi POST) dan 'hasil'
    return render(
        request,
        "index.html",
        {
            "form": form, # Kirim objek form yang sesuai
            "hasil": hasil,
            "hasil_json": json.dumps(hasil),
        }
    )

# ... (di bagian bawah views.py, setelah fungsi 'index')

def analytics_view(request):
    """
    View baru untuk membaca dari file Excel dan menampilkannya di template.
    """
    header = []
    data_rows = []
    error_message = None

    try:
        wb = openpyxl.load_workbook(EVAL_FILE_PATH, read_only=True)
        ws = wb.active

        # Baca semua baris data dari worksheet
        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            error_message = "File log ditemukan tetapi kosong."
        else:
            header = all_rows[0] # Baris pertama adalah header
            data_rows = all_rows[1:] # Sisa baris adalah data
            
            # Balik urutan data agar yang terbaru ada di atas
            data_rows.reverse() 

    except FileNotFoundError:
        error_message = f"File log '{os.path.basename(EVAL_FILE_PATH)}' belum dibuat. Silakan jalankan pencarian rute terlebih dahulu."
    except Exception as e:
        error_message = f"Terjadi kesalahan saat membaca file Excel: {e}"

    return render(request, "analytics.html", {
        "header": header,
        "data_rows": data_rows,
        "error_message": error_message,
    })